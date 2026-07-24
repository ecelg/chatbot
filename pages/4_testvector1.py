import os
import streamlit as st
import pandas as pd
import json
import re
from datetime import datetime

# --- THIRD PARTY EXTRACTION LIBRARIES ---
from pypdf import PdfReader
import docx2txt
import openpyxl

# Pure InterSystems Driver & OpenAI Base APIs
import iris
from openai import OpenAI

# --- PAGE SETUP ---
st.set_page_config(page_title="IRIS Remote Agent Engine", layout="wide")
st.title("🌐 InterSystems IRIS Cloud Knowledge Agent")
st.caption("Deployable on Streamlit.io — Zero hardcoded credentials required.")

# --- ENVIRONMENT SESSION INITIALIZATION ---
if "credentials_loaded" not in st.session_state:
    st.session_state.credentials_loaded = False
if "messages" not in st.session_state:
    st.session_state.messages = []
if "awaiting_web_consent" not in st.session_state:
    st.session_state.awaiting_web_consent = False
if "pending_web_query" not in st.session_state:
    st.session_state.pending_web_query = ""

# --- CONFIGURATION FILE PARSER ---
def parse_config_file(uploaded_file):
    """Parses custom layout credential string matrices directly from memory streams."""
    try:
        content = uploaded_file.read().decode("utf-8")
        lines = content.splitlines()
        config = {}
        for line in lines:
            if "=" in line:
                key, val = line.split("=", 1)
                config[key.strip().lower()] = val.strip()
        
        # Verify presence of critical operating keys
        required_keys = ["str", "port", "namespace", "username", "password", "openaikey"]
        if all(k in config for k in required_keys):
            return config
    except Exception as e:
        st.error(f"Error parsing connection configuration file format: {e}")
    return None

# --- SIDEBAR ACCESS MANAGEMENT ---
with st.sidebar:
    st.header("🔑 Connection Settings")
    st.write("Upload your server credential passkey definition file (.txt) to unlock the application runtime layers.")
    
    config_file = st.file_uploader("Upload credential profile file:", type=["txt"])
    
    if config_file:
        parsed_config = parse_config_file(config_file)
        if parsed_config:
            # Commit keys securely into memory state blocks
            st.session_state.iris_host = parsed_config["str"]
            st.session_state.iris_port = int(parsed_config["port"])
            st.session_state.iris_namespace = parsed_config["namespace"]
            st.session_state.iris_user = parsed_config["username"]
            st.session_state.iris_password = parsed_config["password"]
            st.session_state.openai_key = parsed_config["openaikey"]
            st.session_state.credentials_loaded = True
            st.success("✅ Secure Server Handshake Established!")
        else:
            st.error("❌ Invalid configuration file structure. Please match format schema fields exactly.")
            st.session_state.credentials_loaded = False
            
    if st.session_state.credentials_loaded:
        st.info(f"📍 Connected: {st.session_state.iris_host}:{st.session_state.iris_port}\n🌌 Space: {st.session_state.iris_namespace}")
        if st.button("🔄 Disconnect & Clear Thread Memory"):
            for key in list(st.session_state.keys()):
                del st.session_state[key]
            st.rerun()

# --- BACKEND CONNECTION UTILITIES ---
def get_iris_connection():
    """Connects using the official intersystems-irispython PyPI SDK layout."""
    return iris.connect(
        hostname=st.session_state.iris_host,
        port=st.session_state.iris_port,
        namespace=st.session_state.iris_namespace,
        username=st.session_state.iris_user,
        password=st.session_state.iris_password
    )

def get_openai_client():
    return OpenAI(api_key=st.session_state.openai_key)

# --- UTILITY OPERATIONS ---
def get_embedding(text_string, client=None):
    if client is None:
        client = get_openai_client()
    response = client.embeddings.create(model="text-embedding-3-small", input=[text_string])
    return response.data[0].embedding

def chunk_text(text_data, chunk_size=500, overlap=75):
    words = text_data.split()
    chunks = []
    for i in range(0, len(words), chunk_size - overlap):
        chunk = " ".join(words[i:i + chunk_size])
        if chunk.strip():
            chunks.append(chunk)
    return chunks

def extract_text_from_file(uploaded_file):
    ext = uploaded_file.name.split(".")[-1].lower()
    text_content = ""
    if ext == "pdf":
        for page in PdfReader(uploaded_file).pages:
            text_content += page.extract_text() or ""
    elif ext == "docx":
        text_content = docx2txt.process(uploaded_file)
    elif ext == "xlsx":
        wb = openpyxl.load_workbook(uploaded_file, data_only=True)
        for sheet in wb.sheetnames:
            for row in wb[sheet].iter_rows(values_only=True):
                text_content += " ".join([str(c) for c in row if c is not None]) + "\n"
    return text_content

def log_to_relational_db(log_type, query, status, action):
    """Logs system metrics straight down to the custom remote enterprise tables."""
    try:
        conn = get_iris_connection()
        cursor = conn.cursor()
        cursor.execute("""
            INSERT INTO SQLUser.SystemAdminLogs (LogType, UserQuery, Status, ResolutionAction) 
            VALUES (?, ?, ?, ?)
        """, (log_type, query, status, action))
        conn.commit()
        cursor.close()
        conn.close()
    except:
        pass # Prevents session crashing if admin log tables are not configured on remote target namespaces

def tool_get_database_inventory():
    conn = get_iris_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT Category, FileName, Summary FROM SQLUser.DocumentMetaStore ORDER BY Category")
    rows = cursor.fetchall()
    cursor.close()
    conn.close()
    if not rows:
        return "The database knowledge store is currently completely empty."
    inventory = "Current Database Knowledge Inventory:\n"
    for r in rows:
        inventory += f"- [Category: {r[0]}] File: {r[1]} | Abstract Summary: {r[2]}\n"
    return inventory

def tool_query_vector_db_with_meta(semantic_query, client):
    query_vector = get_embedding(semantic_query, client)
    vector_str = ",".join(map(str, query_vector))
    
    conn = get_iris_connection()
    cursor = conn.cursor()
    
    # Core InterSystems Joint multi-model string collation query
    query = """
        SELECT TOP 3 v.TextChunk, v.SourceFile, m.DocLink, m.Category, 
               VECTOR_COSINE(v.Embedding, TO_VECTOR(?, DOUBLE, 1536)) as Sim
        FROM SQLUser.DocVectors v
        LEFT JOIN SQLUser.DocumentMetaStore m ON %EXACT(v.SourceFile) = %EXACT(m.FileName)
        ORDER BY Sim DESC
    """
    cursor.execute(query, (str(vector_str),))
    results = cursor.fetchall()
    cursor.close()
    conn.close()
    
    THRESHOLD = 0.35
    matched_data = []
    highest_score = 0.0
    
    for r in results:
        try:
            score = float(r[4])
            if score > highest_score:
                highest_score = score
            if score >= THRESHOLD:
                matched_data.append({
                    "text": str(r[0]),
                    "source": str(r[1]),
                    "link": str(r[2]) if r[2] else "",
                    "category": str(r[3]) if r[3] else "General"
                })
        except:
            continue
    return matched_data, highest_score

def tool_get_unique_categories():
    """Queries the remote metadata table to fetch all active categories for the dropdown selection."""
    try:
        conn = get_iris_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT DISTINCT Category FROM SQLUser.DocumentMetaStore WHERE Category IS NOT NULL ORDER BY Category")
        rows = cursor.fetchall()
        cursor.close()
        conn.close()
        return [str(r[0]) for r in rows if r and r[0]]
    except:
        return []

def tool_query_vector_db_filtered(semantic_query, selected_category, client):
    """Performs vector lookups strictly filtered by the user-selected category layer."""
    query_vector = get_embedding(semantic_query, client)
    vector_str = ",".join(map(str, query_vector))
    
    conn = get_iris_connection()
    cursor = conn.cursor()
    
    query = """
        SELECT TOP 3 v.TextChunk, v.SourceFile, m.DocLink, m.Category, 
               VECTOR_COSINE(v.Embedding, TO_VECTOR(?, DOUBLE, 1536)) as Sim
        FROM SQLUser.DocVectors v
        INNER JOIN SQLUser.DocumentMetaStore m ON %EXACT(v.SourceFile) = %EXACT(m.FileName)
        WHERE %EXACT(m.Category) = ?
        ORDER BY Sim DESC
    """
    
    cursor.execute(query, (str(vector_str), str(selected_category)))
    results = cursor.fetchall()
    cursor.close()
    conn.close()
    
    THRESHOLD = 0.35
    matched_data = []
    highest_score = 0.0
    
    for r in results:
        try:
            score = float(r[4])  # Fixed truncated statement from original prompt block
            if score > highest_score:
                highest_score = score
            if score >= THRESHOLD:
                matched_data.append({
                    "text": str(r[0]),
                    "source": str(r[1]),
                    "link": str(r[2]) if r[2] else "",
                    "category": str(r[3]) if r[3] else "General"
                })
        except:
            continue
    return matched_data, highest_score

# --- MAIN SECURITY INTERFACE GATE ---
if not st.session_state.credentials_loaded:
    st.warning("⚠️ Access Restricted: Please upload a valid server connectivity config text file to initiate your session workspace context.")
    st.stop()

# --- INITIALIZE CORE CLIENT ONCE SECURED ---
client = get_openai_client()

# --- NAVIGATION SIDEBAR PANELS ---
st.sidebar.title("Navigation")
page = st.sidebar.radio("Go to:", ["📁 Ingest Documents", "🔍 Browse Knowledge Base", "💬 RAG Knowledge Assistant", "🛡️ Admin Audit Logs"])

# --- PAGE 1: DOCUMENT INGESTION WITH MANUAL CONTROLS ---
if page == "📁 Ingest Documents":
    st.header("Upload & Catalog Context Documents")
    st.write("Upload business documents and manually assign summaries, categories, and references.")
    
    existing_categories = ["--- Select Existing Category ---"]
    try:
        categories = tool_get_unique_categories()
        for cat in categories:
            existing_categories.append(cat)
    except Exception as e:
        pass

    uploaded_file = st.file_uploader("1. Choose a file", type=["pdf", "docx", "xlsx"])
    st.write("### 2. Classification & Metadata")
    
    selected_cat = st.selectbox("Choose a category from your current library:", existing_categories)
    new_cat = st.text_input("OR type a new category name if there is no matching option above:")
    manual_summary = st.text_area("Provide a general summary or overview of this document:", 
                                  placeholder="Type a 2-3 sentence abstract detailing the core context...")
    doc_link = st.text_input("Document Reference Link / URL (Optional):", placeholder="https://sharepoint.com")
    
    st.write("---")
    
    if uploaded_file and st.button("Process & Save Content to IRIS"):
        final_category = ""
        if new_cat.strip():
            final_category = new_cat.strip()
        elif selected_cat != "--- Select Existing Category ---":
            final_category = selected_cat
            
        if not final_category:
            st.error("❌ Please select an existing category or create a new one.")
        elif not manual_summary.strip():
            st.error("❌ Please provide a quick document summary description before continuing.")
        else:
            with st.spinner("Extracting text and calculating vector layouts..."):
                raw_text = extract_text_from_file(uploaded_file)
                
                if not raw_text.strip():
                    st.error("Empty text extracted from the document file container.")
                else:
                    filename = uploaded_file.name
                    doc_id = str(hash(filename) + hash(datetime.now().isoformat()))
                    
                    conn = get_iris_connection()
                    cursor = conn.cursor()
                    
                    cursor.execute("DELETE FROM SQLUser.DocVectors WHERE SourceFile = ?", (filename,))
                    cursor.execute("DELETE FROM SQLUser.DocumentMetaStore WHERE FileName = ?", (filename,))
                    conn.commit()
                    
                    meta_string = json.dumps({
                        "size_bytes": uploaded_file.size, 
                        "timestamp": str(datetime.now()),
                        "source_platform": "Streamlit Form"
                    })
                    
                    cursor.execute("""
                        INSERT INTO SQLUser.DocumentMetaStore (DocID, FileName, Category, Summary, DocLink, DocMetadata) 
                        VALUES (?, ?, ?, ?, ?, ?)
                    """, (doc_id, filename, final_category, manual_summary.strip(), doc_link.strip(), meta_string))
                    
                    chunks = chunk_text(raw_text)
                    for chunk in chunks:
                        vector_array = get_embedding(chunk, client)
                        vector_string = ",".join(map(str, vector_array))
                        cursor.execute("""
                            INSERT INTO SQLUser.DocVectors (SourceFile, TextChunk, Embedding) 
                            VALUES (?, ?, TO_VECTOR(?, DOUBLE, 1536))
                        """, (filename, chunk, vector_string))
                    
                    conn.commit()
                    cursor.close()
                    conn.close()
                    
                    log_to_relational_db("Ingestion", filename, "Success", f"Manually mapped to category: {final_category}")
                    st.success(f"🎉 Successfully ingested '{filename}' under Category: **{final_category}**!")
                    st.rerun()

# --- PAGE 2: BROWSE KNOWLEDGE BASE ---
elif page == "🔍 Browse Knowledge Base":
    st.header("🔍 Browse Document Inventory")
    st.write("Explore metadata and summaries stored inside your InterSystems IRIS cluster relational space.")
    
    try:
        conn = get_iris_connection()
        query = "SELECT Category, FileName, Summary, DocLink FROM SQLUser.DocumentMetaStore ORDER BY Category, FileName"
        df = pd.read_sql(query, conn)
        conn.close()
        
        if df.empty:
            st.info("The document knowledge store is currently completely empty.")
        else:
            categories = df["Category"].unique()
            for cat in categories:
                with st.expander(f"📁 Category: {cat}", expanded=True):
                    cat_df = df[df["Category"] == cat]
                    for _, row in cat_df.iterrows():
                        st.subheader(f"📄 {row['FileName']}")
                        st.write(f"**Summary:** {row['Summary']}")
                        if row['DocLink']:
                            st.markdown(f"🔗 [Reference Link]({row['DocLink']})")
                        st.divider()
    except Exception as e:
        st.error(f"Failed to fetch inventory matrix: {e}")

# --- PAGE 3: RAG KNOWLEDGE ASSISTANT ---
elif page == "💬 RAG Knowledge Assistant":
    st.header("💬 Hybrid Vector RAG Assistant")
    st.write("Query indexed documentation contextual blocks targeted by specific taxonomy groupings.")
    
    # Category targeting filter
    available_cats = tool_get_unique_categories()
    selected_filter = st.selectbox("Target context scoping strictly within category:", ["--- Query Across All Library Items ---"] + available_cats)
    
    # Render chat metrics thread memory logs
    for message in st.session_state.messages:
        with st.chat_message(message["role"]):
            st.markdown(message["content"])
            
    if user_query := st.chat_input("Enter your structural framework question..."):
        with st.chat_message("user"):
            st.markdown(user_query)
        st.session_state.messages.append({"role": "user", "content": user_query})
        
        # Select matching vector collection context
        with st.spinner("Scanning Vector Spaces..."):
            if selected_filter == "--- Query Across All Library Items ---":
                matched_chunks, score = tool_query_vector_db_with_meta(user_query, client)
            else:
                matched_chunks, score = tool_query_vector_db_filtered(user_query, selected_filter, client)
                
        # Build prompt payload context injections
        if matched_chunks:
            context_block = "\n\n".join([f"[Source: {c['source']} | Cat: {c['category']}]: {c['text']}" for c in matched_chunks])
            system_prompt = f"You are a precise corporate knowledge assistant. Answer the user question based strictly on this context map:\n\n{context_block}"
            status_flag, action_log = "Success", f"Context components discovered (Highest Score: {score:.4f})"
        else:
            system_prompt = "You are a corporate knowledge assistant. No matching vector context was discovered; reply notifying the user gracefully."
            context_block = ""
            status_flag, action_log = "No Context Found", "Vector retrieval below operating similarity limits threshold."
            
        # Execute streaming engine via OpenAI
        with st.chat_message("assistant"):
            try:
                response_stream = client.chat.completions.create(
                    model="gpt-4o-mini",
                    messages=[
                        {"role": "system", "content": system_prompt},
                        {"role": "user", "content": user_query}
                    ],
                    stream=True
                )
                assistant_response = st.write_stream(response_stream)
                
                # Append references if found
                if matched_chunks:
                    st.markdown("#### 📚 Referenced Sources:")
                    for c in matched_chunks:
                        link_str = f" ([Link]({c['link']}))" if c['link'] else ""
                        st.caption(f"• **{c['source']}** — *Category: {c['category']}*{link_str}")
            except Exception as e:
                st.error(f"Inference execution failed: {e}")
                assistant_response = f"Error: {e}"
                status_flag = "Failed"
                
        st.session_state.messages.append({"role": "assistant", "content": assistant_response})
        log_to_relational_db("RAG Query", user_query, status_flag, action_log)

# --- PAGE 4: ADMIN AUDIT LOGS ---
elif page == "🛡️ Admin Audit Logs":
    st.header("🛡️ System Administration Audit Ledger Logs")
    st.write("Direct relational ledger queries from InterSystems IRIS operational analytics runtime logs.")
    
    try:
        conn = get_iris_connection()
        query = "SELECT ID, LogType, UserQuery, Status, ResolutionAction FROM SQLUser.SystemAdminLogs ORDER BY ID DESC"
        df = pd.read_sql(query, conn)
        conn.close()
        
        if df.empty:
            st.info("The system audit log table is currently empty.")
        else:
            st.dataframe(df, use_container_width=True, hide_index=True)
    except Exception as e:
        st.error(f"Failed to pull relational log data matrices: {e}")
